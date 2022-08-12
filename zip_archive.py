import os
import os.path
from zipfile import ZipFile
from PyPDF2 import PdfReader
import csv
from openpyxl import load_workbook


def archive_files():
    paths = []
    zip_ = ZipFile('tmp/sample.zip', 'w')
    for filename in os.listdir('resources'):
        paths.append(os.path.join('resources', filename))
        zip_.write(os.path.join('resources', filename))
    return paths


def read_zip_file():
    filenames = []
    zip_1 = archive_files()
    for filename in zip_1:
        if filename.endswith('.pdf') or filename.endswith('.csv') or filename.endswith('.xlsx'):
            filenames.append(filename)
    return filenames


def test_check_zip_files():
    for type_files in read_zip_file():
        if type_files.endswith('.pdf'):
            f = PdfReader(type_files)
            number_of_pages = len(f.pages)
            page = f.pages[0]
            text = page.extract_text()
            assert 1 == number_of_pages
            assert "Пример PDF файла" in text
        elif type_files.endswith('.csv'):
            with open(type_files, newline='') as csvfile:
                spam_reader = csv.reader(csvfile, delimiter=' ', quotechar='|')
                for row in spam_reader:
                    assert 'Number,Footnote' in row
        elif type_files.endswith('.xlsx'):
            workbook = load_workbook(type_files)
            sheet = workbook.active
            assert 'AJ813917' in sheet.cell(row=3, column=2).value
